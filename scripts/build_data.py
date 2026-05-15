#!/usr/bin/env python3
"""
Bake AI_Hedge_Funds_13F_Tracker.xlsx -> data.json for the dashboard.

Run from the repo root:
    python3 scripts/build_data.py [path/to/AI_Hedge_Funds_13F_Tracker.xlsx]

If no path given, defaults to ./AI_Hedge_Funds_13F_Tracker.xlsx
"""
import json
import os
import sys
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# Categorization rules (from the user brief, applied consistently across funds)
# ---------------------------------------------------------------------------
CATEGORIES = {
    "GPU & AI compute":       ["NVDA", "AVGO", "AMD", "ALAB"],
    "Foundry & semis":        ["TSM", "INTC", "TSEM", "ARM"],
    "Semis equipment":        ["AMAT", "ASML", "LRCX", "KLAC"],
    "Memory & storage":       ["MU", "SNDK", "STX", "WDC"],
    "Optics & networking":    ["LITE", "COHR", "ANET", "CIEN"],
    "AI cloud & neoclouds":   ["CRWV", "NBIS", "WYFI"],
    "BTC miners pivoting to AI": [
        "CORZ", "IREN", "APLD", "CIFR", "RIOT", "HUT", "BTDR", "BITF", "CLSK"
    ],
    "Power & energy (data center)": [
        "BE", "VST", "CEG", "GEV", "VRT", "ETN", "EQT", "SEI", "BW",
        "PSIX", "FIX", "EME", "LBRT", "PUMP"
    ],
    "AI software / hyperscaler": [
        "MSFT", "GOOGL", "META", "AMZN", "ORCL", "AAPL", "APP", "CRM"
    ],
}
TICKER_TO_CAT = {t: c for c, tickers in CATEGORIES.items() for t in tickers}


# ---------------------------------------------------------------------------
# Issuer-name → ticker fallback. Some rows in the spreadsheet have a blank
# ticker; we look the issuer up here.
# ---------------------------------------------------------------------------
ISSUER_TO_TICKER = {
    # AI / semis-adjacent
    "Nebius Group N.V.": "NBIS",
    "Lam Research Corp": "LRCX",
    "Advanced Micro Devices Inc": "AMD",
    "Astera Labs Inc": "ALAB",
    "Snowflake Inc": "SNOW",
    "Kla Corp": "KLAC",
    "Asml Holding N V": "ASML",
    "Arm Holdings Plc": "ARM",
    "Marvell Technology Inc": "MRVL",
    "Seagate Technology Hldngs Pl": "STX",
    "Nxp Semiconductors N V": "NXPI",
    "Mks Inc.": "MKSI",
    "Analog Devices Inc": "ADI",
    "Synopsys Inc": "SNPS",
    "Entegris Inc": "ENTG",
    "Qualcomm Inc": "QCOM",
    # Power / data center
    "Talen Energy Corp": "TLN",
    "Argan Inc": "AGX",
    "Nrg Energy Inc": "NRG",
    "Legence Corp": "LGN",
    "Hallador Energy Company": "HNRG",
    "Bkv Corp": "BKV",
    "Tss Inc Del": "TSSI",
    "Aaon Inc": "AAON",
    # Galaxy is a BTC-treasury / data-center hybrid
    "Galaxy Digital Inc.": "GLXY",
    # Mega caps / software
    "Spotify Technology S A": "SPOT",
    "Tesla Inc": "TSLA",
    "Take-Two Interactive Softwar": "TTWO",
    "Apollo Global Mgmt Inc": "APO",
    "Flutter Entmt Plc": "FLUT",
    "Veeva Sys Inc": "VEEV",
    "Corpay Inc": "CPAY",
    "Zillow Group Inc": "Z",
    "Grab Holdings Limited": "GRAB",
    "Block Inc": "SQ",
    "Zscaler Inc": "ZS",
    "Servicenow Inc": "NOW",
    "Liberty Media Corp Del": "FWONK",
    "Workday Inc": "WDAY",
    "Wealthfront Corp": "WLTH",
    "Procore Technologies Inc": "PCOR",
    "Costar Group Inc": "CSGP",
    "Unitedhealth Group Inc": "UNH",
    "Elastic N V": "ESTC",
    "Hinge Health Inc": "HNGE",
    "Sherwin Williams Co": "SHW",
    "Atrenew Inc": "RERE",
    "Pony Ai Inc": "PONY",
    "Circle Internet Group Inc": "CRCL",
    "Jd.Com Inc": "JD",
    "Netskope Inc": "NTSK",
    "Figure Technology Solutio": "FIGR",
    "Klarna Group Plc": "KLAR",
    "Bullish": "BLSH",
    "Etoro Group Ltd": "ETOR",
    "Figma Inc": "FIG",
    "Accelerant Holdings": "ARX",
    "Gemini Space Sta Inc": "GEMI",
    "Mntn Inc": "MNTN",
    "Nu Hldgs Ltd": "NU",
    "Natera Inc": "NTRA",
    "Intuitive Surgical Inc": "ISRG",
    "Adobe Inc": "ADBE",
    "First Ctzns Bancshares Inc D": "FCNCA",
    "Caris Life Sciences Inc": "CAI",
    "Rocket Cos Inc": "RKT",
    "Paypal Hldgs Inc": "PYPL",
    "Axon Enterprise Inc": "AXON",
    "C H Robinson Worldwide Inc": "CHRW",
    "Expeditors Intl Wash Inc": "EXPD",
    "Generac Hldgs Inc": "GNRC",
    "Sprouts Fmrs Mkt Inc": "SFM",
    "Quantumscape Corp": "QS",
    "Chagee Hldgs Ltd": "CHA",
    "S&P Global Inc": "SPGI",
    "Moodys Corp": "MCO",
    "Medical Pptys Trust Inc": "MPW",
    "Navan Inc": "NAVA",
    "Moderna Inc": "MRNA",
    "Ishares Bitcoin Trust Etf": "IBIT",
    "Ishares Tr": "ISHARES",
    "Ishares Inc": "ISHARES",
    "Webull Corp": "BULL",
    "Gitlab Inc": "GTLB",
    "Confluent Inc": "CFLT",
    "Amplitude Inc": "AMPL",
    "Cellebrite Di Ltd": "CLBT",
    "Gds Hldgs Ltd": "GDS",
    "Mastec Inc": "MTZ",
    "Jfrog Ltd": "FROG",
    "Ttm Technologies Inc": "TTMI",
    "Semtech Corp": "SMTC",
    "Fabrinet": "FN",
    "Mongodb Inc": "MDB",
    "Sitime Corp": "SITM",
    "Corning Inc": "GLW",
    "Ciena Corp": "CIEN",
    "Impinj Inc": "PI",
    "Macom Tech Solutions Hldgs I": "MTSI",
    "Klaviyo Inc": "KVYO",
    "Hubspot Inc": "HUBS",
    "Hewlett Packard Enterprise C": "HPE",
    "Okta Inc": "OKTA",
    "Kkr & Co Inc": "KKR",
    "Capital One Finl Corp": "COF",
    "Medline Inc": "MDLN",
    "Vulcan Matls Co": "VMC",
    "Philip Morris Intl Inc": "PM",
    "Carpenter Technology Corp": "CRS",
    "Wingstop Inc": "WING",
    "Amphenol Corp New": "APH",
    "Tenet Healthcare Corp": "THC",
    "Clean Harbors Inc": "CLH",
    "Affirm Hldgs Inc": "AFRM",
    "Mastercard Incorporated": "MA",
    "Transdigm Group Inc": "TDG",
    "Booking Holdings Inc": "BKNG",
    "Boston Scientific Corp": "BSX",
    "Hilton Worldwide Hldgs Inc": "HLT",
    "Visa Inc": "V",
}

# Override category for a handful of issuers where ticker-based mapping is wrong
ISSUER_CATEGORY_OVERRIDE = {
    # The user listed FIG (Figure Technology) under nothing — Figure is a fintech
    # so leave as Other. But "Figma Inc" also resolves to FIG — distinguish.
    # We resolve via issuer name in get_category() below.
}


# ---------------------------------------------------------------------------
# Fund metadata (CIK + filing info). The spreadsheet's Fund Summary tab has
# most of this; CIKs come from the user brief.
# ---------------------------------------------------------------------------
FUND_CIKS = {
    "Situational Awareness LP":           "0002045724",
    "Value Aligned Research Advisors":    "0001963565",
    "Light Street Capital":               "0001569049",
    "Coatue Management":                  "0001135730",
    "Altimeter Capital":                  "0001541617",
    "Tiger Global Management":            "0001167483",
    "Whale Rock Capital Management":      "0001387322",
    "Lone Pine Capital":                  "0001061165",
}

# Map fund-summary names to per-fund tab name
FUND_TAB_BY_NAME = {
    "Situational Awareness LP":           "SA - Aschenbrenner",
    "Value Aligned Research Advisors":    "VARA - Hoskin & Field",
    "Light Street Capital":               "Light Street - Kacher",
    "Coatue Management":                  "Coatue - Laffont",
    "Altimeter Capital":                  "Altimeter - Gerstner",
    "Tiger Global Management":            "Tiger Global - Coleman",
    "Whale Rock Capital Management":      "Whale Rock - Sacerdote",
    "Lone Pine Capital":                  "Lone Pine - Mandel",
}

# Short labels for the UI
FUND_SHORT_LABEL = {
    "Situational Awareness LP":           "Situational Awareness",
    "Value Aligned Research Advisors":    "VARA",
    "Light Street Capital":               "Light Street",
    "Coatue Management":                  "Coatue",
    "Altimeter Capital":                  "Altimeter",
    "Tiger Global Management":            "Tiger Global",
    "Whale Rock Capital Management":      "Whale Rock",
    "Lone Pine Capital":                  "Lone Pine",
}


def get_category(ticker: str, issuer: str) -> str:
    """Return category bucket for a ticker. Issuer is used as a tie-breaker."""
    if ticker in TICKER_TO_CAT:
        return TICKER_TO_CAT[ticker]
    # Override: GLXY is a BTC-treasury / data-center hybrid — bucket as miners
    if ticker == "GLXY":
        return "BTC miners pivoting to AI"
    return "Other"


def resolve_ticker(ticker, issuer):
    """Use ticker if set, otherwise look up issuer."""
    if ticker:
        return ticker.strip().upper()
    if issuer in ISSUER_TO_TICKER:
        return ISSUER_TO_TICKER[issuer]
    # Fallback: build a placeholder from issuer initials
    return "?" + (issuer[:8].upper() if issuer else "UNKNOWN")


def fnum(v):
    """Coerce to float, defaulting to 0."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_fund_summary(wb):
    """Read the Fund Summary tab and return dict of fund metadata."""
    ws = wb["Fund Summary"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    funds = {}
    for r in rows[1:]:
        if not r or r[1] is None:
            continue
        name = r[1]
        filing_date = r[10]
        if isinstance(filing_date, datetime):
            filing_date = filing_date.strftime("%Y-%m-%d")
        funds[name] = {
            "name": name,
            "shortLabel": FUND_SHORT_LABEL.get(name, name),
            "manager": r[2],
            "cik": FUND_CIKS.get(name),
            "filesOwn13F": r[1] in FUND_CIKS,
            "totalBn": r[3],
            "numIssuers": r[4],
            "equityPct": r[5],
            "callsPct": r[6],
            "putsPct": r[7],
            "optionsHeavy": (r[8] == "YES"),
            "top10Concentration": r[9],
            "filingDate": filing_date if isinstance(filing_date, str) else None,
            "thesis": r[11],
            "rank": r[0] if isinstance(r[0], int) else None,
            "isReference": r[0] == "REF",
        }
    return funds


def parse_fund_holdings(wb, tab_name):
    """Read a per-fund tab and return list of holdings."""
    ws = wb[tab_name]
    rows = list(ws.iter_rows(values_only=True))
    holdings = []
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == "Rank" and r[1] == "Ticker":
            header_idx = i
            break
    if header_idx is None:
        return holdings
    for r in rows[header_idx + 1:]:
        if not r or not isinstance(r[0], int):
            continue
        rank = r[0]
        raw_ticker = r[1]
        issuer = (r[2] or "").strip()
        equity = fnum(r[3])
        calls = fnum(r[4])
        puts = fnum(r[5])
        total = fnum(r[6])
        pct = fnum(r[7])
        ticker = resolve_ticker(raw_ticker, issuer)
        category = get_category(ticker, issuer)
        holdings.append({
            "rank": rank,
            "ticker": ticker,
            "issuer": issuer,
            "equity": equity,
            "calls": calls,
            "puts": puts,
            "total": total,
            "pct": pct,
            "category": category,
            "tickerInSheet": raw_ticker is not None,
        })
    return holdings


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "AI_Hedge_Funds_13F_Tracker.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True)

    funds_meta = parse_fund_summary(wb)
    funds_out = []
    for name, meta in funds_meta.items():
        tab = FUND_TAB_BY_NAME.get(name)
        if tab and tab in wb.sheetnames:
            meta["holdings"] = parse_fund_holdings(wb, tab)
            meta["hasHoldings"] = True
            meta["periodOfReport"] = "2025-12-31"  # spreadsheet period
            meta["sourceTag"] = "xlsx"
        else:
            meta["holdings"] = []
            meta["hasHoldings"] = False
        funds_out.append(meta)

    # Period info (from brief)
    out = {
        "asOf": "2025-12-31",
        "builtAt": datetime.utcnow().isoformat() + "Z",
        "source": "SEC EDGAR 13F filings, period Dec 31 2025",
        "categories": list(CATEGORIES.keys()) + ["Other"],
        "categoryColors": {
            "GPU & AI compute":            "#76b900",  # nvidia green
            "Foundry & semis":             "#5b8def",
            "Semis equipment":             "#9b7cf2",
            "Memory & storage":            "#f4a261",
            "Optics & networking":         "#06d6a0",
            "AI cloud & neoclouds":        "#48cae4",
            "BTC miners pivoting to AI":   "#f4a300",
            "Power & energy (data center)":"#e76f51",
            "AI software / hyperscaler":   "#bdb2ff",
            "Other":                       "#5a5a5a",
        },
        "funds": funds_out,
        "caveats": [
            "13F shows US-listed longs only — no shorts, no foreign listings (SK Hynix, Samsung, Kioxia not visible), no privates.",
            "Options reported at underlying notional, not premium paid — so options-heavy funds (SA, VARA) have inflated 13F totals vs. capital deployed.",
            "Whale Rock AI Fund and Point72 Turion don't file standalone 13Fs — they roll up into parent filings.",
        ],
    }

    out_path = "data.json"
    if len(sys.argv) > 2:
        out_path = sys.argv[2]
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {out_path}: {len(funds_out)} funds, {sum(len(f['holdings']) for f in funds_out)} holdings")

    # Also bake the JSON straight into index.html so the page works via file://
    # (double-clicking the html) without needing a local HTTP server.
    html_path = "index.html"
    if os.path.exists(html_path):
        with open(html_path) as f:
            html = f.read()
        # Compact JSON, with </ escaped so a stray </script> in the data
        # can't break out of the inline <script> block.
        inline = json.dumps(out, separators=(",", ":")).replace("</", "<\\/")
        import re
        new_html = re.sub(
            r'(<script id="bakedData" type="application/json">)([\s\S]*?)(</script>)',
            lambda m: m.group(1) + inline + m.group(3),
            html, count=1,
        )
        if new_html != html:
            with open(html_path, "w") as f:
                f.write(new_html)
            print(f"Updated inline data in {html_path}")

    # Print category coverage so we can sanity-check
    cat_counts = {}
    for f in funds_out:
        for h in f["holdings"]:
            cat_counts[h["category"]] = cat_counts.get(h["category"], 0) + 1
    print("Category coverage:")
    for c, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")


if __name__ == "__main__":
    main()
